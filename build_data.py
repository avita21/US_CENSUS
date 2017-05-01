#!/usr/bin/env python

# used to get the US Census key
import os
# import libraries for API requests
try:
    # For Python 3.0 and later
    from urllib.request import urlopen, Request, URLError
except ImportError:
    # Fall back to Python 2's urllib2
    from urllib2 import Request, urlopen, URLError

# import libraries for printing prettily
import pprint
# for converting data
import pandas as pd
import numpy as np

# REMEMBER TO ADD YOUR CENSUS KEY TO THE SYSTEM VARIABLES!!!!!
key = os.environ['USCENSUS_KEY']; 

# make a request from the US Census:
def census_request(url):
    request = Request(url+'&key='+key);
    try:
        response = urlopen(request);
        data     = response.read();
        df       = pd.read_json(data); 
        return df;
    except (URLError, e):
        print('Got an error code:', e);
        raise;


# format the education dataset:
def education():
    # get the (Bachelor's Degree or higher) Education Rate from 2010-2014 
    # for each state, or:
    # EDUCATIONAL ATTAINMENT!!Population 25 years and over!!Percent bachelor's
    # degree or higher
    
    # 2010 and 2011 are in CSV files:
    df = pd.DataFrame();
    for (path, year) in [('Datasets/ACS_10_5YR_S1501/ACS_10_5YR_S1501_with_ann.csv',
                          '2010'),
                         ('Datasets/ACS_11_5YR_S1501/ACS_11_5YR_S1501_with_ann.csv',
                          '2011')]:
        next_year = pd.read_csv(path, header=None, names=["state", year], 
                                usecols=[1, 3], skiprows=2, dtype={'state':str});
        next_year = next_year.set_index('state');
        df = pd.concat([df, next_year], axis=1); 
    
    # 2012-2014 are from the Census API:
    for i in [2012, 2013, 2014]:
        acs_education_URL = 'https://api.census.gov/data/'+str(i)+ \
                            '/acs5/profile?get=DP02_0067PE&for=state:*';
        next_year = census_request(acs_education_URL);
        next_year = next_year.rename(columns = {0:str(i), 1:'state'});
        next_year.drop(0, inplace=True);
        next_year = next_year.set_index('state');
        df = pd.concat([df, next_year], axis=1); 
    return df;

# format the job creation dataset
def job_creation():
    # get the Job Creation Rate from 2010-2014 for each state:
    bds_URL = 'https://api.census.gov/data/bds/firms?get=' + \
              'job_creation_rate_births,sic1&for=state:*&time=from+2010+to+2014';
    bds_jobcreation = census_request(bds_URL);
    return bds_jobcreation;

# Build datasets:
def main():
    Jobs    = job_creation();
    Edu     = education();
    Uni     = pd.read_excel('Datasets/University_States.xlsx', 
                            sheetname='HERD2015_DST_62', header=3, skip_footer=4);    
    MinWage = pd.read_table('Datasets/min_wage.txt', sep='\t', header=2); 
    GDP     = pd.read_excel('Datasets/gdplev.xls', header=None, 
                            skiprows=np.arange(0,89,1),
                            skip_footer=195, names=['Years', 'GDP in Billions'], 
                            parse_cols='A:B');
    Elec    = pd.read_csv('Datasets/Average_retail_price_of_electricity.csv',
                          header=None, names=["State", "2010","2011","2012",
                                              "2013","2014"],
                          index_col = 0, usecols=[0, 3, 4, 5, 6, 7],
                          skiprows=6);
    Unemp   = pd.read_excel('Datasets/annual_unemployment.xlsx', 
                            header=None, skiprows=np.arange(0,7,1),
                            skip_footer=2, names=["State", "2010","2011","2012",
                                                    "2013","2014"], 
                            parse_cols='A,T,Q,N,K,H');
    
if __name__ == "__main__":
    main();

'''
## Code Used for Converting XLS Excel Docs to XSLX for use with openpyxl
## From: Ray, at: https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in xrange(0, nrows):
        for col in xrange(0, ncols):
            print "r: " + str(row) + " c: " + str(col)
            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1
'''
